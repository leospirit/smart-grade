from fastapi import FastAPI, Depends, UploadFile, File, HTTPException, status, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from typing import List
from pydantic import BaseModel
import pandas as pd
import io
import json

import models, schemas, crud, auth
from database import SessionLocal, engine, get_db
import uuid
import os

models.Base.metadata.create_all(bind=engine)

app = FastAPI(title="Smart Grade Platform")

# Create default admin on startup
@app.on_event("startup")
def create_default_admin():
    db = SessionLocal()
    try:
        admin = db.query(models.User).filter(models.User.username == "admin").first()
        if not admin:
            hashed_pwd = auth.get_password_hash("admin123")
            admin_user = models.User(username="admin", hashed_password=hashed_pwd, role="admin")
            db.add(admin_user)
            db.commit()
            print("Default admin created: admin/admin123")
    finally:
        db.close()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # For dev only, relax to *
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Auth Endpoints
@app.post("/api/token", response_model=schemas.Token)
async def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.username == form_data.username).first()
    if not user or not auth.verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token = auth.create_access_token(data={"sub": user.username})
    return {"access_token": access_token, "token_type": "bearer"}

@app.get("/api/users/me", response_model=schemas.User)
async def read_users_me(current_user: models.User = Depends(auth.get_current_active_user)):
    return current_user



@app.post("/api/upload/roster")
async def upload_roster(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Upload a master roster Excel file.
    Expected columns: '学号' (Student Number), '姓名' (Name), '班级' (Class)
    """
    contents = await file.read()
    df = pd.read_excel(io.BytesIO(contents))
    
    # Normalize headers
    # Simple mapping for demo purposes
    df.columns = [str(c).strip() for c in df.columns]
    
    # Map columns
    col_map = {}
    for c in df.columns:
        if "学号" in c or "ID" in c.upper(): col_map["id"] = c
        elif "姓名" in c or "Name" in c.upper(): col_map["name"] = c
        elif "班级" in c or "Class" in c.upper(): col_map["class"] = c
        elif "年级" in c or "Grade" in c.upper(): col_map["grade"] = c
        
    count = 0
    for index, row in df.iterrows():
        s_id = str(row[col_map["id"]]) if "id" in col_map else f"unknown_{index}"
        name = str(row[col_map["name"]]) if "name" in col_map else "Unknown"
        c_name = str(row[col_map["class"]]) if "class" in col_map else "Default Class"
        g_name = str(row[col_map["grade"]]) if "grade" in col_map else "Default Grade"
        
        # Check exist
        existing = crud.get_student_by_number(db, s_id)
        if not existing:
            # Create a StudentCreate schema object
            student_data = schemas.StudentCreate(
                student_number=s_id,
                name=name,
                class_name=c_name,
                grade_name=g_name # Assuming schemas.StudentCreate and models.Student have grade_name
            )
            new_student = crud.create_student(db, student_data)
            
            # Create User Account for Student
            # Check if user already exists (unlikely if student is new, but safecheck)
            user_exists = db.query(models.User).filter(models.User.username == s_id).first()
            if not user_exists:
                hashed_pwd = auth.get_password_hash("123456")
                new_user = models.User(
                    username=s_id, 
                    hashed_password=hashed_pwd, 
                    role="student", 
                    student_id=new_student.id
                )
                db.add(new_user)
            
            count += 1
        else:
            # Update info if needed
            existing.name = name
            existing.class_name = c_name
            existing.grade_name = g_name # Update grade_name
            db.add(existing) # Re-add to session for update
    
    db.commit()
    return {"message": f"Successfully imported {count} new students."}

@app.post("/api/upload/grades")
async def upload_grades(course_name: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Smart upload for grades.
    1. Auto-detects header row (looks for '学号', 'ID', '姓名', 'Name').
    2. Finds 'Total' column.
    3. Treat all other columns as sub-scores.
    """
    try:
        contents = await file.read()
        # Read without header first to find the real header row
        df_raw = pd.read_excel(io.BytesIO(contents), header=None)
        
        # Find header row
        header_row_idx = 0
        found_header = False
        
        # Scan first 10 rows
        for i in range(min(10, len(df_raw))):
            row_values = df_raw.iloc[i].astype(str).tolist()
            # Check for keywords
            if any("学号" in v or "ID" in v.upper() for v in row_values) and \
               any("姓名" in v or "Name" in v.upper() for v in row_values):
                header_row_idx = i
                found_header = True
                break
        
        if found_header:
            # Reload with correct header
            df = pd.read_excel(io.BytesIO(contents), header=header_row_idx)
        else:
            # Fallback to default
            df = pd.read_excel(io.BytesIO(contents))

        # Ensure course exists
        course = crud.get_course_by_name(db, course_name)
        if not course:
            course = crud.create_course(db, course_name)
        
        # Identify Key Columns
        student_col = None
        name_col = None
        total_col = None
        
        for col in df.columns:
            c_str = str(col).strip()
            # ID
            if any(k in c_str for k in ['学号', 'Student ID', 'Student No', '学籍号']):
                student_col = col
            elif c_str.upper() == 'ID': # Strict for just 'ID'
                student_col = col
            
            # Name
            elif any(k in c_str for k in ['姓名', 'Name', 'Student Name']):
                name_col = col
            
            # Total
            elif any(k in c_str for k in ['总分', '成绩', '得分', 'Total', 'Score']):
                total_col = col

        matched_count = 0
        unmatched = []

        if not student_col and not name_col:
             raise HTTPException(status_code=400, detail="Could not find '学号' or '姓名' columns in Excel.")

        for _, row in df.iterrows():
            student = None
            # Prioritize ID
            if student_col:
                val = str(row[student_col]).strip()
                # Remove .0 if it's a float-string (e.g. "1001.0")
                if val.endswith('.0'): val = val[:-2]
                student = crud.get_student_by_number(db, val)
            
            # Fallback to Name
            if not student and name_col:
                 val = str(row[name_col]).strip()
                 # Simple match (Warning: duplicates)
                 student = db.query(models.Student).filter(models.Student.name == val).first()
            
            if not student:
                unmatched.append(str(row.to_dict()))
                continue
                
            # Parse Total Score
            total_score = 0.0
            if total_col:
                try:
                    total_score = float(row[total_col])
                except:
                    pass
            
            # Sub-scores (Everything else)
            sub_scores = {}
            for col in df.columns:
                if col not in [student_col, name_col, total_col]:
                    # Ignore empty/unnamed columns
                    if "Unnamed" in str(col): continue
                    
                    val = row[col]
                    try: 
                        if pd.isna(val): continue
                        # Clean key name
                        key_name = str(col).strip()
                        sub_scores[key_name] = val
                        
                        # Calculate total if 0 (optional feature)
                        # if total_score == 0 and isinstance(val, (int, float)):
                        #    total_score += val
                    except:
                        pass
                        
            crud.create_or_update_grade(db, student.id, course.id, total_score, sub_scores)
            matched_count += 1
            
        return {
            "message": f"Processed grades for {course_name}",
            "matched": matched_count,
            "unmatched_count": len(unmatched),
            "unmatched_rows": unmatched[:5] # Return top 5 errors
        }

    except Exception as e:
        print(f"Upload Error: {e}")
        raise HTTPException(status_code=500, detail=f"Server Error: {str(e)}")

class ImportConfirmRequest(BaseModel):
    file_key: str
    course_name: str
    mapping: dict

@app.post("/api/upload/preview")
async def upload_preview(file: UploadFile = File(...)):
    """
    Step 1: Save file and return columns for mapping.
    """
    if not os.path.exists("temp"):
        os.makedirs("temp")
        
    file_ext = file.filename.split(".")[-1]
    file_key = f"{uuid.uuid4()}.{file_ext}"
    file_path = f"temp/{file_key}"
    
    with open(file_path, "wb") as f:
        f.write(await file.read())
    
    # Smart Scan Header
    df_raw = pd.read_excel(file_path, header=None)
    header_idx = 0
    
    # Scan first 10 rows for keywords
    for i in range(min(10, len(df_raw))):
        row_values = df_raw.iloc[i].astype(str).tolist()
        if any("学号" in v or "ID" in v.upper() for v in row_values) or \
           any("姓名" in v or "Name" in v.upper() for v in row_values):
            header_idx = i
            break
            
    # Read columns
    df = pd.read_excel(file_path, header=header_idx)
    # Get first few rows for preview
    # Handle NaN in preview to avoid JSON error
    df_preview = df.head(3).fillna('')
    preview = df_preview.astype(str).to_dict(orient='records')
    
    return {
        "file_key": file_key,
        "columns": list(df.columns),
        "preview": preview,
        "detected_header_row": header_idx
    }

@app.post("/api/upload/confirm")
async def upload_confirm(req: ImportConfirmRequest, db: Session = Depends(get_db)):
    """
    Step 2: Process file with user-defined mapping.
    """
    file_path = f"temp/{req.file_key}"
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File expired or not found. Please upload again.")
        
    try:
        # Re-read file (we assume header row detection was correct or we could pass it too, 
        # but reusing the smart scan logic is redundant. Ideally pass header_idx. 
        # For MVP, we scan again or trust default/smart scan works same way.
        # Let's Scan again to be safe.
        df_raw = pd.read_excel(file_path, header=None)
        header_idx = 0
        for i in range(min(10, len(df_raw))):
            row_values = df_raw.iloc[i].astype(str).tolist()
            if any("学号" in v or "ID" in v.upper() for v in row_values) or \
               any("姓名" in v or "Name" in v.upper() for v in row_values):
                header_idx = i
                break
        
        df = pd.read_excel(file_path, header=header_idx)
        
        # Ensure course
        course = crud.get_course_by_name(db, req.course_name)
        if not course:
            course = crud.create_course(db, req.course_name)
            
        mapping = req.mapping
        student_col = mapping.get("student_id")
        name_col = mapping.get("name")
        total_col = mapping.get("total_score")
        
        matched_count = 0
        
        for _, row in df.iterrows():
            student = None
            if student_col and student_col in df.columns:
                val = str(row[student_col]).strip()
                if val.endswith('.0'): val = val[:-2]
                student = crud.get_student_by_number(db, val)
                
            if not student and name_col and name_col in df.columns:
                val = str(row[name_col]).strip()
                student = db.query(models.Student).filter(models.Student.name == val).first()
                
            if not student:
                 continue
                 
            # Total Score
            total_score = 0.0
            if total_col and total_col in df.columns:
                try:
                    total_score = float(row[total_col])
                except:
                    pass
            
            # Sub Scores (All columns NOT selected in mapping)
            sub_scores = {}
            used_cols = [student_col, name_col, total_col]
            for col in df.columns:
                if col not in used_cols:
                     if "Unnamed" in str(col): continue
                     val = row[col]
                     if pd.isna(val): continue
                     sub_scores[str(col).strip()] = val
            
            crud.create_or_update_grade(db, student.id, course.id, total_score, sub_scores)
            matched_count += 1
            
        # Cleanup
        os.remove(file_path)
        
        return {"message": f"Successfully imported {matched_count} records.", "matched": matched_count}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/students")
def read_students(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    students = crud.get_all_students(db, skip=skip, limit=limit)
    # Enrich with grades for simple view
    result = []
    for s in students:
        s_dict = {
            "id": s.id,
            "student_number": s.student_number,
            "name": s.name,
            "class_name": s.class_name,
            "grades": {}
        }
        for g in s.grades:
            s_dict["grades"][g.course.name] = {
                "total": g.total_score,
                "details": g.sub_scores
            }
        result.append(s_dict)
    return result

@app.get("/api/export/roster")
def export_roster(class_name: str = None, db: Session = Depends(get_db)):
    """
    Export roster to Excel.
    Locks the 'Student ID' column (Column A) to prevent editing.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Protection
    from fastapi.responses import StreamingResponse

    # 1. Fetch Students
    query = db.query(models.Student)
    if class_name:
        query = query.filter(models.Student.class_name == class_name)
    students = query.all()

    # 2. Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Roster"

    # Header
    headers = ["学号", "姓名", "班级", "总分"] # Keep it simple for score entry
    ws.append(headers)

    # Data
    for s in students:
        ws.append([s.student_number, s.name, s.class_name, ""])

    # 3. Apply Protection
    # By default, openpyxl locks all cells if worksheet protection is on.
    # We want to UNLOCK everything EXCEPT Column A.
    
    # Unlock all cells first
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=False)
    
    # Lock Column A (Student ID) explicitly
    # Iterating valid rows to lock A column
    for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
        for cell in row:
            cell.protection = Protection(locked=True)
            
    # Also lock header row? Maybe better to keep it editable just in case, 
    # but strictly speaking we want IDs fixed. Let's start with just ID column cells.

    ws.protection.sheet = True # Enable sheet protection
    ws.protection.password = "123456" # Simple password to prevent accidental edits
    ws.protection.enable()

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"roster_{class_name if class_name else 'all'}.xlsx"
    
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    
    return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.delete("/api/courses/{course_name}")
def delete_course(course_name: str, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admin can delete courses")
    students = db.query(models.Student).all()
    count = 0
    for s in students:
        if s.grades and course_name in s.grades:
            # Create a copy to trigger SQLAlchemy update detection for JSON
            new_grades = dict(s.grades)
            del new_grades[course_name]
            s.grades = new_grades
            db.add(s) # Mark as modified
            count += 1
    db.commit()
    # Also delete from Course table
    crud.delete_course_by_name(db, course_name)
    return {"message": f"Course '{course_name}' deleted from {count} students."}

@app.get("/api/courses")
def read_courses(db: Session = Depends(get_db)):
    """Get all courses with visibility status"""
    courses = db.query(models.Course).all()
    # If a course exists in DB but not in student grades, it shows up here.
    # ALSO, we might want to include courses that are only in student grades but not in Course table? 
    # Current upload_grades ensures Course entity exists.
    return [{"name": c.name, "is_visible": c.is_visible} for c in courses]

@app.put("/api/courses/{course_name}/toggle")
def toggle_course_visibility(course_name: str, db: Session = Depends(get_db)):
    course = crud.get_course_by_name(db, course_name)
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    
    course.is_visible = not course.is_visible
    db.commit()
    return {"message": f"Course '{course_name}' visibility set to {course.is_visible}", "is_visible": course.is_visible}

# --- User Management Endpoints ---

@app.get("/api/users", response_model=List[schemas.User])
async def read_users(skip: int = 0, limit: int = 100, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admin can view users")
    return crud.get_users(db, skip=skip, limit=limit)

@app.post("/api/users", response_model=schemas.User)
async def create_user(user: schemas.UserCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admin can create users")
    db_user = crud.get_user_by_username(db, username=user.username)
    if db_user:
        raise HTTPException(status_code=400, detail="Username already registered")
    return crud.create_user(db=db, user=user)

@app.delete("/api/users/{user_id}")
async def delete_user(user_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admin can delete users")
    if current_user.id == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    if not crud.delete_user(db, user_id):
         raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deleted"}

@app.post("/api/users/{user_id}/reset-password")
async def reset_password(user_id: int, request: schemas.PasswordReset, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admin can reset passwords")
    
    updated_user = crud.reset_user_password(db, user_id, request.new_password)
    if not updated_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "Password reset successfully"}

@app.post("/api/users/me/password")
async def change_password(request: schemas.PasswordChange, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    if current_user.role in ["parent", "student"]:
        raise HTTPException(status_code=403, detail="Parents/Students cannot change password")
    
    # Verify old password
    if not auth.verify_password(request.old_password, current_user.hashed_password):
        raise HTTPException(status_code=400, detail="Incorrect old password")
    
    updated_user = crud.update_user_password(db, current_user.id, request.new_password)
    return {"message": "Password changed successfully"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
